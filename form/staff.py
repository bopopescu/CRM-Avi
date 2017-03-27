from os import getcwd, path, mkdir, listdir, rmdir
from shutil import copy
from form.templates import list
from PyQt5.uic import loadUiType
from PyQt5.QtWidgets import QDialog, QMainWindow, QMessageBox, QTableWidgetItem, QListWidgetItem, QFileDialog, QLineEdit, QWidget, QSizePolicy
from PyQt5.QtGui import QIcon
from PyQt5.QtCore import Qt, QDate
from function import my_sql, to_excel
import openpyxl
from openpyxl.drawing.image import Image
import subprocess
from classes.my_class import User


staff_list_class = loadUiType(getcwd() + '/ui/staff.ui')[0]
one_staff_class = loadUiType(getcwd() + '/ui/add_work.ui')[0]
country_class = loadUiType(getcwd() + '/ui/country.ui')[0]
position_class = loadUiType(getcwd() + '/ui/staff_position.ui')[0]
exel_info_class = loadUiType(getcwd() + '/ui/exel_info.ui')[0]
info_date_class = loadUiType(getcwd() + '/ui/exel_info_date.ui')[0]
add_file_date_class = loadUiType(getcwd() + '/ui/work_add_file.ui')[0]
staff_filter = loadUiType(getcwd() + '/ui/staff_filter.ui')[0]


class Staff(QMainWindow, staff_list_class):
    def __init__(self, main=None, dc_select=False):
        super(Staff, self).__init__()
        self.setupUi(self)
        self.setWindowIcon(QIcon(getcwd() + "/images/icon.ico"))
        self.main = main
        self.dc_select = dc_select
        self.filter = None
        self.query_table_all = """SELECT staff_worker_info.Id, First_Name, Last_Name, DATE_FORMAT(Date_Recruitment, '%d.%m.%Y'), `Leave`, Date_Leave, staff_position.Name
                                    FROM staff_worker_info LEFT JOIN staff_position ON staff_worker_info.Position_Id = staff_position.Id"""
        self.query_table_select = self.query_table_all

        self.set_settings()
        self.set_info()

        # Быстрый фильтр
        self.le_fast_filter = QLineEdit()
        self.le_fast_filter.setPlaceholderText("Фамилия")
        self.le_fast_filter.setMaximumWidth(150)
        self.le_fast_filter.editingFinished.connect(self.fast_filter)
        dummy = QWidget()
        dummy.setSizePolicy(QSizePolicy.MinimumExpanding, QSizePolicy.Preferred)
        self.toolBar.addWidget(dummy)
        self.toolBar.addWidget(self.le_fast_filter)
        self.access()

    def access(self):
        for item in User().access_list(self.__class__.__name__):
            a = getattr(self, item["atr1"])
            if item["atr2"]:
                a = getattr(a, item["atr2"])

            if item["value"]:
                a(item["value"])
            else:
                a()

    def inspection_path(self, dir_name, sql_dir_name):  # Находим путь работника
        if not hasattr(self, 'path_work'):
            query = 'SELECT `Values` FROM program_settings_path WHERE Name = "%s"' % sql_dir_name
            info_sql = my_sql.sql_select(query)
            if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql", info_sql.msg, QMessageBox.Ok)
                        return False
            self.path_wor = info_sql[0][0]
            if not path.isdir("%s/%s" % (self.path_wor, dir_name)):
                try:
                    mkdir("%s/%s" % (self.path_wor, dir_name))
                    return "%s/%s" % (self.path_wor, dir_name)
                except:
                    QMessageBox.critical(self, "Ошибка файлы", "Нет доступа к корневому диалогу, файлы недоступны", QMessageBox.Ok)
                    return False
            else:
                return "%s/%s" % (self.path_wor, dir_name)

    def set_settings(self):
        self.tw_workers.horizontalHeader().resizeSection(0, 50)
        self.tw_workers.horizontalHeader().resizeSection(1, 150)
        self.tw_workers.horizontalHeader().resizeSection(2, 150)
        self.tw_workers.horizontalHeader().resizeSection(3, 80)

    def set_info(self):
        query = "SELECT Id, Name FROM staff_position"
        self.staff_positions = my_sql.sql_select(query)
        if "mysql.connector.errors" in str(type(self.staff_positions)):
            QMessageBox.critical(self, "Ошибка sql", self.staff_positions.msg, QMessageBox.Ok)
            return False
        self.staff_workers = my_sql.sql_select(self.query_table_select)
        if "mysql.connector.errors" in str(type(self.staff_workers)):
            QMessageBox.critical(self, "Ошибка sql", self.staff_workers.msg, QMessageBox.Ok)
            return False

        self.lw_position.clear()
        for position in self.staff_positions:
            self.lw_position.addItem(position[1])
        else:
            self.lw_position.addItem("Уволеные в этом месяце")
            self.lw_position.addItem("Уволеные в этом году")
            self.lw_position.addItem("Уволеные")
            self.lw_position.addItem("Все работающие")

        self.tw_workers.clearContents()
        self.tw_workers.setRowCount(len(self.staff_workers))
        for row in range(len(self.staff_workers)):
            for column in range(4):
                a = self.staff_workers[row][column]
                item = QTableWidgetItem(str(a))
                self.tw_workers.setItem(row, column, item)

    def add(self):
        self.add_mat = OneStaff(self)
        self.add_mat.set_add_settings()
        self.add_mat.setWindowModality(Qt.ApplicationModal)
        self.add_mat.show()

    def double_click(self, row):
        if not self.dc_select:
            id = self.tw_workers.item(row, 0).text()
            self.add_mat = OneStaff(self, True)
            self.add_mat.set_add_settings()
            if self.add_mat.insert_info(id):
                self.add_mat.setWindowModality(Qt.ApplicationModal)
                self.add_mat.show()
        else:
            item = (self.tw_workers.item(row, 0).text(), self.tw_workers.item(row, 2).text())
            self.main.of_list_worker(item)
            self.close()
            self.destroy()

    def change(self):
        try:
            row = self.tw_workers.selectedItems()[0].row()
            self.double_click(row)
        except:
            pass

    def delete(self):
        try:
            select_work = self.tw_workers.item(self.tw_workers.selectedItems()[0].row(), 0).text()
            if select_work:
                result = QMessageBox.question(self, "Удаление", "Точно удалить работника?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                if result == 16384:
                    query = "SELECT Last_Name, First_Name, DATE_FORMAT(Date_Recruitment, '%d.%m.%Y') FROM staff_worker_info WHERE Id = %s"
                    dir_name = my_sql.sql_select(query, (select_work, ))
                    if "mysql.connector.errors" in str(type(dir_name)):
                        QMessageBox.critical(self, "Ошибка sql", dir_name.msg, QMessageBox.Ok)
                        return False
                    query = "DELETE FROM staff_worker_info WHERE Id = %s"
                    info_sql = my_sql.sql_change(query, (select_work, ))
                    if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql", info_sql.msg, QMessageBox.Ok)
                        return False
                    dir = dir_name[0][0] + " " + dir_name[0][1] + " " + dir_name[0][2]
                    rmdir(self.inspection_path(dir, "Путь корень рабочие"))
                    self.set_info()
        except:
            pass

    def sorting(self, select_position):
        self.select_position = select_position  # Запоменаем выбраную должность
        self.to_date = QDate.currentDate()

        if self.select_position == "Уволеные":  # Вставляем уволеных
            self.tw_workers.clearContents()
            self.tw_workers.setRowCount(0)
            for row in range(len(self.staff_workers)):
                if self.staff_workers[row][4] == 1:
                    self.tw_workers.insertRow(self.tw_workers.rowCount())
                    for column in range(4):
                        a = self.staff_workers[row][column]
                        item = QTableWidgetItem(str(a))
                        self.tw_workers.setItem(self.tw_workers.rowCount() - 1, column, item)

        elif self.select_position == "Уволеные в этом году":  # Вставляем уволеных
            self.tw_workers.clearContents()
            self.tw_workers.setRowCount(0)
            for row in range(len(self.staff_workers)):
                if self.staff_workers[row][4] == 1 and self.staff_workers[row][5].year == self.to_date.year():
                    self.tw_workers.insertRow(self.tw_workers.rowCount())
                    for column in range(4):
                        a = self.staff_workers[row][column]
                        item = QTableWidgetItem(str(a))
                        self.tw_workers.setItem(self.tw_workers.rowCount() - 1, column, item)

        elif self.select_position == "Уволеные в этом месяце":  # Вставляем уволеных
            self.tw_workers.clearContents()
            self.tw_workers.setRowCount(0)
            for row in range(len(self.staff_workers)):
                if self.staff_workers[row][4] == 1 and self.staff_workers[row][5].month == self.to_date.month() and self.staff_workers[row][5].year == self.to_date.year():
                    self.tw_workers.insertRow(self.tw_workers.rowCount())
                    for column in range(4):
                        a = self.staff_workers[row][column]
                        item = QTableWidgetItem(str(a))
                        self.tw_workers.setItem(self.tw_workers.rowCount() - 1, column, item)

        elif self.select_position == "Все работающие":  # Вставляем уволеных
            self.tw_workers.clearContents()
            self.tw_workers.setRowCount(0)
            for row in range(len(self.staff_workers)):
                if self.staff_workers[row][4] == 0:
                    self.tw_workers.insertRow(self.tw_workers.rowCount())
                    for column in range(4):
                        a = self.staff_workers[row][column]
                        item = QTableWidgetItem(str(a))
                        self.tw_workers.setItem(self.tw_workers.rowCount() - 1, column, item)

        else:
            self.tw_workers.clearContents()
            self.tw_workers.setRowCount(0)
            for row in range(len(self.staff_workers)):
                if self.staff_workers[row][6] == self.select_position and self.staff_workers[row][4] == 0:
                    self.tw_workers.insertRow(self.tw_workers.rowCount())
                    for column in range(4):
                        a = self.staff_workers[row][column]
                        item = QTableWidgetItem(str(a))
                        self.tw_workers.setItem(self.tw_workers.rowCount() - 1, column, item)

    def ui_filter(self):
        if self.filter is None:
            self.filter = StaffFilter(self)
        self.filter.of_set_sql_query(self.query_table_all)
        self.filter.setWindowModality(Qt.ApplicationModal)
        self.filter.show()

    def fast_filter(self):
        # Блок условий артикула
        if self.le_fast_filter.text() != '':
            q_filter = "(staff_worker_info.Last_Name LIKE '%s')" % ("%" + self.le_fast_filter.text() + "%", )
            self.query_table_select = self.query_table_all + " WHERE " + q_filter
        else:
            self.query_table_select = self.query_table_all

        self.set_info()

    def of_set_filter(self, sql):
        self.query_table_select = sql

        self.set_info()


class OneStaff(QMainWindow, one_staff_class):
    def __init__(self, main, change=False):
        super(OneStaff, self).__init__()
        self.setupUi(self)
        self.setWindowIcon(QIcon(getcwd() + "/images/icon.ico"))
        self.change = change  # Запоминаем это добаление работника или изменение
        self.m = main
        self.alert = []  # Массив для запоминания изменений
        self.access()

    def access(self):
        for item in User().access_list(self.__class__.__name__):
            a = getattr(self, item["atr1"])
            if item["atr2"]:
                a = getattr(a, item["atr2"])

            if item["value"]:
                a(item["value"])
            else:
                a()

    def inspection_path(self, dir_name, sql_dir_name):  # Находим путь работника
        if not hasattr(self, 'path_work'):
            query = 'SELECT `Values` FROM program_settings_path WHERE Name = "%s"' % sql_dir_name
            info_sql = my_sql.sql_select(query)
            if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql", info_sql.msg, QMessageBox.Ok)
                        return False
            self.path_wor = info_sql[0][0]
            if not path.isdir("%s/%s" % (self.path_wor, dir_name)):
                try:
                    mkdir("%s/%s" % (self.path_wor, dir_name))
                    return "%s/%s" % (self.path_wor, dir_name)
                except:
                    QMessageBox.critical(self, "Ошибка файлы", "Нет доступа к корневому диалогу, файлы недоступны", QMessageBox.Ok)
                    return False
            else:
                return "%s/%s" % (self.path_wor, dir_name)

    def inspection_files(self, dir_name, sql_dir_name):   # Проверяем файлы и даем иконки
        self.path = self.inspection_path(dir_name, sql_dir_name)
        if self.path:
            self.lw_file.clear()
            files = listdir("%s/%s" % (self.path_wor, dir_name))
            for file in files:
                if "~" not in file:
                    r = path.splitext(file)  # Получаем название и расширение
                    if "xlsx" in r[1][1:] or "xlsm" in r[1] or "xls" in r[1] or "xlt" in r[1]:
                        ico = "xlsx"
                    elif "xml" in r[1][1:] or "docx" in r[1] or "doc" in r[1] or "docm" in r[1]:
                        ico = "xml"
                    elif "png" in r[1].lower() or "jpg" in r[1] or "jpeg" in r[1] or "jpe" in r[1] or "gif" in r[1] or "bmp" in r[1]:
                        ico = "image"
                    elif "pdf" in r[1]:
                        ico = "pdf"
                    else:
                        ico = "other"

                    list_item = QListWidgetItem(r[0] + r[1])
                    list_item.setIcon(QIcon(getcwd() + "/images/%s.ico" % ico))
                    self.lw_file.addItem(list_item)

    def set_add_settings(self):
        # Начальные чеки
        self.de_info_leave.setEnabled(False)
        self.rb_sex_f.setChecked(True)

        # Выставляем даты
        self.to_date = QDate.currentDate()
        self.de_info_birth.setDate(self.to_date)
        self.de_info_recruitment.setDate(self.to_date)
        self.de_info_leave.setDate(self.to_date)
        self.de_passport_issued.setDate(self.to_date)
        self.de_passport_ending.setDate(self.to_date)
        self.de_migration_validity_from.setDate(self.to_date)
        self.de_migration_validity_to.setDate(self.to_date)
        self.de_registration_validity_from.setDate(self.to_date)
        self.de_registration_validity_to.setDate(self.to_date)
        self.de_registration.setDate(self.to_date)
        self.de_patent_issued.setDate(self.to_date)
        self.de_patent_ending.setDate(self.to_date)
        self.de_migration.setDate(self.to_date)
        self.de_insurance_date.setDate(self.to_date)
        self.de_notification.setDate(self.to_date)

        # заполняем страны
        query = "SELECT Country_name FROM staff_country"
        self.country = my_sql.sql_select(query)
        if "mysql.connector.errors" in str(type(self.country)):
            QMessageBox.critical(self, "Ошибка sql", self.country.msg, QMessageBox.Ok)
            return False
        for country in self.country:
            self.cb_info_country.addItem(country[0])

        # заполняем должности
        query = "SELECT Name FROM staff_position"
        self.position = my_sql.sql_select(query)
        if "mysql.connector.errors" in str(type(self.position)):
            QMessageBox.critical(self, "Ошибка sql", self.position.msg, QMessageBox.Ok)
            return False
        for country in self.position:
            self.cb_info_position.addItem(country[0])

        self.alert = []  # Обнуляем массив для запоминания изменений

    def change_birth(self, birth_date):  # Подсчет колличества лет
        years = int(birth_date.daysTo(self.to_date) / 365)
        self.lb_info_years.setText("Возраст: %s" % years)

    def change_patent_date(self, patent_date):  # автоматически продлевает патент на 1 год
        self.de_patent_ending.setDate(patent_date.addYears(1))

    def select_file(self, file):  # Открываем выбраный фаил
        dir_name = self.le_info_last_name.text() + " " + self.le_info_first_name.text() + " " + self.de_info_recruitment.date().toString("dd.MM.yyyy")
        self.path = self.inspection_path(dir_name, 'Путь корень рабочие')
        if self.path:
            file_name = file.text()
            subprocess.Popen(r'%s/%s' % (self.path.replace("/", "\\"), file_name.replace("/", "\\")), shell=True)

    def open_dir(self):  # Открываем выбраную папку
        dir_name = self.le_info_last_name.text() + " " + self.le_info_first_name.text() + " " + self.de_info_recruitment.date().toString("dd.MM.yyyy")
        self.path = self.inspection_path(dir_name, 'Путь корень рабочие')
        if self.path:
            # subprocess.Popen(['explorer "' + self.path.replace("/", "\\") + '"'])
            subprocess.Popen('explorer "%s"' % self.path.replace("/", "\\"))

    def add_file(self):  # Добавляем файлы
        info = AddFile()
        if info.exec() == 0:
            return False
        new_r = path.splitext(info.path_copy_file.text())[1]
        dir_name = self.le_info_last_name.text() + " " + self.le_info_first_name.text() + " " + self.de_info_recruitment.date().toString("dd.MM.yyyy")
        copy(info.path_copy_file.text(), self.inspection_path(dir_name, 'Путь корень рабочие') + "/" + info.le_new_file_name.text() + path.splitext(info.path_copy_file.text())[1])
        self.inspection_files(dir_name, 'Путь корень рабочие')

    def input_check(self):
        self.alert2 = []  # Запоминаем введеные даные
        self.delete = []  # Запоминаем что надо удалить
        # проверяем основную информацию
        if self.le_info_first_name.text() == "":
            QMessageBox.critical(self, "Ошибка ввода", "Вы не ввели имя", QMessageBox.Ok)
            return False
        if self.le_info_last_name.text() == "":
            QMessageBox.critical(self, "Ошибка ввода", "Вы не ввели фамилию", QMessageBox.Ok)
            return False
        if self.de_info_birth.date() > self.to_date:
            QMessageBox.critical(self, "Ошибка ввода", "Не верная дата рождения", QMessageBox.Ok)
            return False
        if self.cb_info_leave.isChecked() and self.de_info_leave.date() < self.de_info_recruitment.date():
            QMessageBox.critical(self, "Ошибка ввода", "Не верная дата увольнения", QMessageBox.Ok)
            return False
        if self.le_info_birthplace.text() == "":
            QMessageBox.critical(self, "Ошибка ввода", "Вы не ввели местро рождения", QMessageBox.Ok)
            return False

        # проверка паспорта
        if self.le_passport_number.text() != "" or self.le_passport_issued.text() != "":
            self.alert2.append("passport")
            if self.le_passport_number.text() == "":
                QMessageBox.critical(self, "Ошибка ввода", "Вы не ввели номер паспорта", QMessageBox.Ok)
                return False
            if self.le_passport_issued.text() == "":
                QMessageBox.critical(self, "Ошибка ввода", "Вы не ввели кем выдан пасспорт", QMessageBox.Ok)
                return False
            if self.de_passport_ending.date() < self.de_passport_issued.date():
                QMessageBox.critical(self, "Ошибка ввода", "Не верная дата окончания паспорта", QMessageBox.Ok)
                return False
        elif self.le_passport_series.text() == "" and self.le_passport_number.text() == "" and self.le_passport_issued.text() == "":
            self.delete.append("passport")
            self.alert2.append("passport")

        # проверка миграционной карты
        if self.le_migration_serial.text() != "" or self.le_migration_number.text() != "":
            self.alert2.append("migration")
            if self.le_migration_serial.text() == "":
                QMessageBox.critical(self, "Ошибка ввода", "Вы не ввели серию миграционной карты", QMessageBox.Ok)
                return False
            if self.le_migration_number.text() == "":
                QMessageBox.critical(self, "Ошибка ввода", "Вы не ввели номер миграционной карты", QMessageBox.Ok)
                return False
            if self.le_migration_kpp.text() == "":
                QMessageBox.critical(self, "Ошибка ввода", "Вы не ввели КПП выдачи минрационной карты", QMessageBox.Ok)
                return False
            if self.de_migration_validity_to.date() < self.de_migration_validity_from.date():
                QMessageBox.critical(self, "Ошибка ввода", "Не верная дата окончания миграционной карты", QMessageBox.Ok)
                return False
        elif self.le_migration_serial.text() == "" and self.le_migration_number.text() == "" and self.le_migration_kpp.text() == "":
            self.delete.append("migration")
            self.alert2.append("migration")

        # проверка регистрации
        if self.le_registration_address.text() != "":
            self.alert2.append("registration")
            if self.de_registration_validity_to.date() < self.de_registration_validity_from.date():
                QMessageBox.critical(self, "Ошибка ввода", "Не верная дата окончания регистрации", QMessageBox.Ok)
                return False
        elif self.le_registration_address.text() == "":
            self.delete.append("registration")
            self.alert2.append("registration")

        # проверка патента
        if self.le_patent_serial.text() != "" or self.le_patent_number.text() != "" or self.le_patent_additional_number.text() != "":
            self.alert2.append("patent")
            if self.le_patent_serial.text() == "":
                QMessageBox.critical(self, "Ошибка ввода", "Вы не ввели серию патента", QMessageBox.Ok)
                return False
            if self.le_patent_number.text() == "":
                QMessageBox.critical(self, "Ошибка ввода", "Вы не ввели номер патента", QMessageBox.Ok)
                return False
            if self.le_patent_additional_number.text() == "":
                QMessageBox.critical(self, "Ошибка ввода", "Вы не ввели дополнительный номер патета", QMessageBox.Ok)
                return False
            if self.le_patent_issued.text() == "":
                QMessageBox.critical(self, "Ошибка ввода", "Вы не ввели кем выдан патент", QMessageBox.Ok)
                return False
            if self.de_patent_ending.date() < self.de_patent_issued.date():
                QMessageBox.critical(self, "Ошибка ввода", "Не верная дата окончания патента", QMessageBox.Ok)
                return False
        elif self.le_patent_serial.text() == "" and self.le_patent_number.text() == "" and self.le_patent_additional_number.text() == "" and self.le_patent_issued.text() == "":
            self.delete.append("patent")
            self.alert2.append("patent")

        # проверка страховки
        if self.le_insurance_number.text() != "" or self.le_insurance_company.text() != "" :
            self.alert2.append("insurance")
            if self.le_insurance_number.text() == "":
                QMessageBox.critical(self, "Ошибка ввода", "Вы не ввели номер страховой", QMessageBox.Ok)
                return False
            if self.le_insurance_company.text() == "":
                QMessageBox.critical(self, "Ошибка ввода", "Вы не ввели компанию страховой", QMessageBox.Ok)
                return False
        elif self.le_insurance_number.text() == "" and self.le_insurance_company.text() == "":
            self.delete.append("insurance")
            self.alert2.append("insurance")

        if self.le_login_login.text() == "" and self.le_login_password.text() == "":
            self.delete.append("login")

        return True

    def alter_info(self):  # Инфрмация об изменении
        if "info" not in self.alert:
            self.alert.append("info")

    def alter_passport(self):  # Инфрмация об изменении
        if "passport" not in self.alert:
            self.alert.append("passport")

    def alter_migration(self):  # Инфрмация об изменении
        if "migration" not in self.alert:
            self.alert.append("migration")

    def alter_registration(self):  # Инфрмация об изменении
        if "registration" not in self.alert:
            self.alert.append("registration")

    def alter_patent(self):  # Инфрмация об изменении
        if "patent" not in self.alert:
            self.alert.append("patent")

    def alter_insurance(self):  # Инфрмация об изменении
        if "insurance" not in self.alert:
            self.alert.append("insurance")

    def alter_notification(self):  # Инфрмация об изменении
        if "notification" not in self.alert:
            self.alert.append("notification")

    def alter_login(self):  # Инфрмация об изменении
        if "login" not in self.alert:
            self.alert.append("login")

    def insert_info(self, id_worker):
        # Обнуляем список заполненых полей
        self.update_sql = []
        # Вставляем основную информацию
        self.id_info = id_worker
        query = """SELECT staff_worker_info.Id, First_Name, Last_Name, Middle_Name, Sex, Date_Birth, Date_Recruitment, staff_worker_info.Leave,
                    Date_Leave, staff_country.Country_name, Phone, Address, staff_position.Name, INN, SNILS, Note, Birthplace
                    FROM staff_worker_info LEFT JOIN staff_country ON staff_worker_info.Country_Id = staff_country.Id
                    LEFT JOIN staff_position ON staff_worker_info.Position_Id = staff_position.Id  WHERE staff_worker_info.Id = %s"""
        sql_reply = my_sql.sql_select(query, (id_worker,))
        if "mysql.connector.errors" in str(type(sql_reply)):
            QMessageBox.critical(self, "Ошибка sql", sql_reply.msg, QMessageBox.Ok)
            return False
        self.lb_info_id.setText("Табельный номер: %s" % sql_reply[0][0])
        self.le_info_first_name.setText(sql_reply[0][1])
        self.le_info_last_name.setText(sql_reply[0][2])
        self.le_info_middle_name.setText(sql_reply[0][3])
        if sql_reply[0][4] == "M":
            self.rb_sex_m.setChecked(True)
        elif sql_reply[0][4] == "F":
            self.rb_sex_f.setChecked(True)
        self.le_info_birthplace.setText(sql_reply[0][16])
        self.de_info_birth.setDate(sql_reply[0][5])
        self.de_info_recruitment.setDate(sql_reply[0][6])
        if sql_reply[0][7]:
            self.cb_info_leave.setChecked(True)
            self.de_info_leave.setEnabled(True)
        self.de_info_leave.setDate(sql_reply[0][8])
        self.cb_info_country.setCurrentText(sql_reply[0][9])
        self.le_info_phone.setText(sql_reply[0][10])
        self.le_info_address.setText(sql_reply[0][11])
        self.cb_info_position.setCurrentText(sql_reply[0][12])
        self.le_info_inn.setText(sql_reply[0][13])
        self.le_info_snils.setText(sql_reply[0][14])
        self.le_info_note.appendPlainText(sql_reply[0][15])
        leave = sql_reply[0][7]

        # Заполняем паспорт
        query = "SELECT Series, Number, Issued, Data_Issued, Date_Ending FROM staff_worker_passport WHERE Worker_Info_Id = %s"
        sql_reply = my_sql.sql_select(query, (id_worker,))
        if "mysql.connector.errors" in str(type(sql_reply)):
            QMessageBox.critical(self, "Ошибка sql", sql_reply.msg, QMessageBox.Ok)
            return False
        if sql_reply:
            self.update_sql.append("passport")
            self.le_passport_series.setText(sql_reply[0][0])
            self.le_passport_number.setText(sql_reply[0][1])
            self.le_passport_issued.setText(sql_reply[0][2])
            self.de_passport_issued.setDate(sql_reply[0][3])
            self.de_passport_ending.setDate(sql_reply[0][4])

        # Заполняем миграционку
        query = "SELECT Serial, Number, KPP, Date_Validity_From, Date_Validity_To, Date_migration FROM staff_worker_migration WHERE Worker_Info_Id = %s"
        sql_reply = my_sql.sql_select(query, (id_worker,))
        if "mysql.connector.errors" in str(type(sql_reply)):
            QMessageBox.critical(self, "Ошибка sql", sql_reply.msg, QMessageBox.Ok)
            return False
        if sql_reply:
            self.update_sql.append("migration")
            self.le_migration_serial.setText(sql_reply[0][0])
            self.le_migration_number.setText(sql_reply[0][1])
            self.le_migration_kpp.setText(sql_reply[0][2])
            self.de_migration_validity_from.setDate(sql_reply[0][3])
            self.de_migration_validity_to.setDate(sql_reply[0][4])
            self.de_migration.setDate(sql_reply[0][5])

        # Заполняем регистрацию
        query = "SELECT Address, Date_Registration, Date_Validity_From, Date_Validity_To FROM staff_worker_registraton WHERE Worker_Info_id = %s"
        sql_reply = my_sql.sql_select(query, (id_worker,))
        if "mysql.connector.errors" in str(type(sql_reply)):
            QMessageBox.critical(self, "Ошибка sql", sql_reply.msg, QMessageBox.Ok)
            return False
        if sql_reply:
            self.update_sql.append("registration")
            self.le_registration_address.setText(sql_reply[0][0])
            self.de_registration.setDate(sql_reply[0][1])
            self.de_registration_validity_from.setDate(sql_reply[0][2])
            self.de_registration_validity_to.setDate(sql_reply[0][3])

        # Заполняем патент
        query = "SELECT Serial, Number, Additional_Number, Issued, Data_Issued, Date_Ending FROM staff_worker_patent WHERE Worker_Info_Id = %s"
        sql_reply = my_sql.sql_select(query, (id_worker,))
        if "mysql.connector.errors" in str(type(sql_reply)):
            QMessageBox.critical(self, "Ошибка sql", sql_reply.msg, QMessageBox.Ok)
            return False
        if sql_reply:
            self.update_sql.append("patent")
            self.le_patent_serial.setText(sql_reply[0][0])
            self.le_patent_number.setText(sql_reply[0][1])
            self.le_patent_additional_number.setText(sql_reply[0][2])
            self.le_patent_issued.setText(sql_reply[0][3])
            self.de_patent_issued.setDate(sql_reply[0][4])
            self.de_patent_ending.setDate(sql_reply[0][5])

        # Заполняем страховку
        query = "SELECT Number, Company, Date FROM staff_worker_insurance WHERE Worker_Info_Id = %s"
        sql_reply = my_sql.sql_select(query, (id_worker,))
        if "mysql.connector.errors" in str(type(sql_reply)):
            QMessageBox.critical(self, "Ошибка sql", sql_reply.msg, QMessageBox.Ok)
            return False
        if sql_reply:
            self.update_sql.append("insurance")
            self.le_insurance_number.setText(sql_reply[0][0])
            self.le_insurance_company.setText(sql_reply[0][1])
            self.de_insurance_date.setDate(sql_reply[0][2])

        # Заполняем Оповещения
        query = "SELECT Date, Note FROM staff_worker_notification WHERE Worker_Info_Id = %s"
        sql_reply = my_sql.sql_select(query, (id_worker,))
        if "mysql.connector.errors" in str(type(sql_reply)):
            QMessageBox.critical(self, "Ошибка sql", sql_reply.msg, QMessageBox.Ok)
            return False
        if sql_reply:
            self.update_sql.append("notification")
            self.cb_notification.setChecked(True)
            self.le_notofication.setEnabled(True)
            self.de_notification.setDate(sql_reply[0][0])
            self.le_insurance_number.setText(sql_reply[0][1])

        # Заполняем список файлов
        dir_name = self.le_info_last_name.text() + " " + self.le_info_first_name.text() + " " + self.de_info_recruitment.date().toString("dd.MM.yyyy")
        self.inspection_files(dir_name, 'Путь корень рабочие')

        # Заполняем логин
        if leave == 0:
            query = "SELECT Login, Password FROM staff_worker_login WHERE Worker_Info_Id = %s"
            sql_reply = my_sql.sql_select(query, (id_worker,))
            if "mysql.connector.errors" in str(type(sql_reply)):
                QMessageBox.critical(self, "Ошибка sql", sql_reply.msg, QMessageBox.Ok)
                return False
            if sql_reply:
                self.update_sql.append("login")
                self.le_login_login.setText(sql_reply[0][0])
                self.le_login_password.setText(sql_reply[0][1])

        self.alert = []
        return True

    def check_login(self):
        query = "SELECT COUNT(*) FROM staff_worker_login WHERE Login = %s"
        info_sql = my_sql.sql_select(query, (self.le_login_login.text(), ))
        if "mysql.connector.errors" in str(type(info_sql)):
            QMessageBox.critical(self, "Ошибка sql", info_sql.msg, QMessageBox.Ok)
            return False
        if info_sql[0][0] > 0:
            QMessageBox.critical(self, "Занято", "Этот логин занят", QMessageBox.Ok)
        else:
            QMessageBox.information(self, "Свободно", "Этот логин свободен", QMessageBox.Ok)

    def free_login(self):
        query = """SELECT (staff_worker_login.Login+1) as `login`
                    FROM staff_worker_login
                    WHERE ( SELECT 1 FROM staff_worker_login as `st` WHERE `st`.Login = (staff_worker_login.Login + 1) ) IS NULL
                    ORDER BY staff_worker_login.Login LIMIT 1"""
        info_sql = my_sql.sql_select(query)
        if "mysql.connector.errors" in str(type(info_sql)):
            QMessageBox.critical(self, "Ошибка sql", info_sql.msg, QMessageBox.Ok)
            return False

        text = "Логин: %s свободен" % str(int(info_sql[0][0]))
        QMessageBox.information(self, "Логин", text, QMessageBox.Ok)

    def acc(self):  # Добаление информации в базу
        if self.change:  # Если мы изменяем а не добавляем работника
            if self.input_check():  # Проверка заполнености полей
                if "info" in self.alert:  # Добаление основной информации
                    if self.rb_sex_m.isChecked():  # Узнаем пол работника
                        self.sex = "M"
                    elif self.rb_sex_f.isChecked():
                        self.sex = "F"
                    id_country = my_sql.sql_select("SELECT Id FROM staff_country WHERE Country_name = %s", (self.cb_info_country.currentText(),))[0][0]
                    id_position = my_sql.sql_select("SELECT Id FROM staff_position WHERE Name = %s", (self.cb_info_position.currentText(),))[0][0]

                    query = """UPDATE staff_worker_info SET First_Name = %s, Last_Name = %s, Middle_Name = %s, Sex = %s, Date_Birth = %s, Birthplace = %s, Date_Recruitment = %s,
                        `Leave` = %s,  Date_Leave = %s, Country_Id = %s, Phone = %s, Address = %s, Position_Id = %s, INN = %s, SNILS = %s, Note = %s WHERE Id = %s"""
                    parametrs = (self.le_info_first_name.text(), self.le_info_last_name.text(), self.le_info_middle_name.text(), self.sex,
                                 self.de_info_birth.date().toString(Qt.ISODate), self.le_info_birthplace.text(),
                                 self.de_info_recruitment.date().toString(Qt.ISODate),
                                 self.cb_info_leave.isChecked(), self.de_info_leave.date().toString(Qt.ISODate), id_country, self.le_info_phone.text(),
                                 self.le_info_address.text(), id_position, self.le_info_inn.text(), self.le_info_snils.text(), self.le_info_note.toPlainText(),
                                 self.id_info)
                    info_sql = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql i", info_sql.msg, QMessageBox.Ok)
                        return False

                if "passport" in self.alert and "passport" in self.alert2:
                    if "passport" in self.delete:  # Если пспорт надо удалить
                        query = "DELETE FROM staff_worker_passport WHERE Worker_Info_Id = %s"
                        parametrs = (self.id_info, )
                    elif "passport" in self.update_sql:  # Если пспорт надо обновить
                        query = "UPDATE staff_worker_passport SET Series = %s, Number = %s, Issued = %s, Data_Issued = %s, Date_Ending = %s WHERE Worker_Info_Id = %s"
                        parametrs = (self.le_passport_series.text(), self.le_passport_number.text(), self.le_passport_issued.text(),
                                     self.de_passport_issued.date().toString(Qt.ISODate), self.de_passport_ending.date().toString(Qt.ISODate), self.id_info)
                    else:  # Если пспорт надо добавить
                        query = "INSERT INTO staff_worker_passport (Worker_Info_Id, Series, Number, Issued, Data_Issued, Date_Ending) VALUES (%s, %s, %s, %s, %s, %s)"
                        parametrs = (self.id_info, self.le_passport_series.text(), self.le_passport_number.text(), self.le_passport_issued.text(),
                                     self.de_passport_issued.date().toString(Qt.ISODate), self.de_passport_ending.date().toString(Qt.ISODate))
                    info_sql = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql p", info_sql.msg, QMessageBox.Ok)
                        return False

                if "migration" in self.alert and "migration" in self.alert2:
                    if "migration" in self.delete:  # Если пспорт надо удалить
                        query = "DELETE FROM staff_worker_migration WHERE Worker_Info_Id = %s"
                        parametrs = (self.id_info, )
                    elif "migration" in self.update_sql:  # Если Миграцию надо обновить
                        query = """UPDATE staff_worker_migration SET Serial = %s, Number = %s, KPP = %s, Date_migration = %s, Date_Validity_From = %s,
                                 Date_Validity_To = %s WHERE Worker_Info_Id = %s"""
                        parametrs = (self.le_migration_serial.text(), self.le_migration_number.text(), self.le_migration_kpp.text(),
                                     self.de_migration.date().toString(Qt.ISODate),
                                     self.de_migration_validity_from.date().toString(Qt.ISODate), self.de_migration_validity_to.date().toString(Qt.ISODate),
                                     self.id_info)
                    else:  # Если миграцию надо добавить
                        query = "INSERT INTO staff_worker_migration (Worker_Info_Id, Serial, Number, KPP, Date_migration, Date_Validity_From, Date_Validity_To) VALUES (%s, %s, %s, %s, %s, %s, %s)"
                        parametrs = (self.id_info, self.le_migration_serial.text(), self.le_migration_number.text(), self.le_migration_kpp.text(),
                                     self.de_migration.date().toString(Qt.ISODate), self.de_migration_validity_from.date().toString(Qt.ISODate),
                                     self.de_migration_validity_to.date().toString(Qt.ISODate))
                    info_sql = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql m", info_sql.msg, QMessageBox.Ok)
                        return False

                if "registration" in self.alert and "registration" in self.alert2:
                    if "registration" in self.delete:  # Если пспорт надо удалить
                        query = "DELETE FROM staff_worker_registraton WHERE Worker_Info_Id = %s"
                        parametrs = (self.id_info, )
                    elif "registration" in self.update_sql:  # Если регистрацию надо обновить
                        query = "UPDATE staff_worker_registraton SET Address = %s, Date_Registration = %s, Date_Validity_From = %s, Date_Validity_To = %s WHERE Worker_Info_id = %s"
                        parametrs = (self.le_registration_address.text(), self.de_registration.date().toString(Qt.ISODate),
                                     self.de_registration_validity_from.date().toString(Qt.ISODate), self.de_registration_validity_to.date().toString(Qt.ISODate),
                                     self.id_info)
                    else:  # Если регистрацию надо добавить
                        query = "INSERT INTO staff_worker_registraton (Worker_Info_id, Address, Date_Registration, Date_Validity_From, Date_Validity_To) VALUES (%s, %s, %s, %s, %s)"
                        parametrs = (self.id_info, self.le_registration_address.text(), self.de_registration.date().toString(Qt.ISODate),
                                     self.de_registration_validity_from.date().toString(Qt.ISODate), self.de_registration_validity_to.date().toString(Qt.ISODate))
                    info_sql = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql r", info_sql.msg, QMessageBox.Ok)
                        return False

                if "patent" in self.alert and "patent" in self.alert2:
                    if "patent" in self.delete:  # Если пспорт надо удалить
                        query = "DELETE FROM staff_worker_patent WHERE Worker_Info_Id = %s"
                        parametrs = (self.id_info, )
                    elif "patent" in self.update_sql:  # Если патент надо обновить
                        query = """UPDATE staff_worker_patent SET Serial = %s, Number = %s, Additional_Number = %s, Issued = %s, Data_Issued = %s, Date_Ending = %s
                                   WHERE Worker_Info_Id = %s"""
                        parametrs = (self.le_patent_serial.text(), self.le_patent_number.text(), self.le_patent_additional_number.text(),
                                     self.le_patent_issued.text(), self.de_patent_issued.date().toString(Qt.ISODate),
                                     self.de_patent_ending.date().toString(Qt.ISODate), self.id_info)
                    else:  # Если патент надо добавить
                        query = "INSERT INTO staff_worker_patent (Worker_Info_Id, Serial, Number, Additional_Number, Issued,  Data_Issued, Date_Ending) VALUES (%s, %s, %s, %s, %s, %s, %s)"
                        parametrs = (self.id_info, self.le_patent_serial.text(), self.le_patent_number.text(), self.le_patent_additional_number.text(),
                                     self.le_patent_issued.text(), self.de_patent_issued.date().toString(Qt.ISODate),
                                     self.de_patent_ending.date().toString(Qt.ISODate))
                    info_sql = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql pt", info_sql.msg, QMessageBox.Ok)
                        return False

                if "insurance" in self.alert and "insurance" in self.alert2:
                    if "insurance" in self.delete:  # Если пспорт надо удалить
                        query = "DELETE FROM staff_worker_insurance WHERE Worker_Info_Id = %s"
                        parametrs = (self.id_info, )
                    elif "insurance" in self.update_sql:  # Если страховку надо обновить
                        query = """UPDATE staff_worker_insurance SET Number = %s, Company = %s, Date = %s WHERE Worker_Info_Id = %s"""
                        parametrs = (self.le_insurance_number.text(), self.le_insurance_company.text(),
                                     self.de_insurance_date.date().toString(Qt.ISODate), self.id_info)
                    else:  # Если страховку надо добавить
                        query = "INSERT INTO staff_worker_insurance (Worker_Info_Id, Number, Company, Date) VALUES (%s, %s, %s, %s)"
                        parametrs = (self.id_info, self.le_insurance_number.text(), self.le_insurance_company.text(), self.de_insurance_date.date().toString(Qt.ISODate))
                    info_sql = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql is", info_sql.msg, QMessageBox.Ok)
                        return False

                if "notification" in self.alert:
                    if "notification" in self.update_sql:  # Если напоминания надо обновить
                        if self.cb_notification.isChecked():
                            query = """UPDATE staff_worker_notification SET Date = %s, Note = %s WHERE Worker_Info_Id = %s"""
                            parametrs = (self.de_notification.date().toString(Qt.ISODate, self.le_notofication.toPlainText()), self.id_info)
                        else:
                            query = """DELETE FROM staff_worker_notification WHERE Worker_Info_Id = %s"""
                            parametrs = (self.id_info, )
                    else:  # Если напоминания надо добавить
                        if self.cb_notification.isChecked():
                            query = "INSERT INTO staff_worker_notification (Worker_Info_Id, Note, Date) VALUES (%s, %s, %s)"
                            parametrs = (self.id_info, self.le_notofication.toPlainText(), self.de_notification.date().toString(Qt.ISODate))
                    info_sql = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql nt", info_sql.msg, QMessageBox.Ok)
                        return False

                if "login" in self.update_sql:  # Если логин надо обновить
                    if not self.cb_info_leave.isChecked() and "login" in self.alert and "login" not in self.delete:
                        query = """UPDATE staff_worker_login SET Login = %s, Password = %s WHERE Worker_Info_Id = %s"""
                        parametrs = (self.le_login_login.text(), self.le_login_password.text(), self.id_info)
                    else:
                        query = """DELETE FROM staff_worker_login WHERE Worker_Info_Id = %s"""
                        parametrs = (self.id_info, )
                elif not self.cb_info_leave.isChecked():  # Если логин надо добавить
                    query = "INSERT INTO staff_worker_login (Worker_Info_Id, Login, Password) VALUES (%s, %s, %s)"
                    parametrs = (self.id_info, self.le_login_login.text(), self.le_login_password.text())
                try:
                    info_sql = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(info_sql)):
                        if info_sql.errno == 1062:
                            QMessageBox.critical(self, "Ошибка sql", "Такой логин уже есть он не сохраниться!", QMessageBox.Ok)
                            return False
                        else:
                            QMessageBox.critical(self, "Ошибка sql lg", info_sql.msg, QMessageBox.Ok)
                            return False
                except:
                    pass
                self.alert = []  # Обнуляем массив для запоминания изменений
                self.close()
                self.m.set_info()
                self.destroy()

        else:  # Если мы добавляем работника
            if self.input_check():  # Проверка заполнености полей
                if "info" in self.alert:  # Добаление основной информации
                    if self.rb_sex_m.isChecked():  # Узнаем пол работника
                        self.sex = "M"
                    elif self.rb_sex_f.isChecked():
                        self.sex = "F"
                    id_country = my_sql.sql_select("SELECT Id FROM staff_country WHERE Country_name = %s", (self.cb_info_country.currentText(),))[0][0]
                    id_position = my_sql.sql_select("SELECT Id FROM staff_position WHERE Name = %s", (self.cb_info_position.currentText(),))[0][0]

                    query = """INSERT INTO staff_worker_info (First_Name, Last_Name, Middle_Name, Sex, Date_Birth, Birthplace, Date_Recruitment, `Leave`,
                                                              Date_Leave, Country_Id, Phone, Address, Position_Id, INN, SNILS, Note)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
                    parametrs = (self.le_info_first_name.text(), self.le_info_last_name.text(), self.le_info_middle_name.text(), self.sex,
                                 self.de_info_birth.date().toString(Qt.ISODate), self.le_info_birthplace.text(),
                                 self.de_info_recruitment.date().toString(Qt.ISODate),
                                 self.cb_info_leave.isChecked(), self.de_info_leave.date().toString(Qt.ISODate), id_country, self.le_info_phone.text(),
                                 self.le_info_address.text(), id_position, self.le_info_inn.text(), self.le_info_snils.text(), self.le_info_note.toPlainText())
                    self.id_info = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(self.id_info)):
                        QMessageBox.critical(self, "Ошибка sql", self.id_info.msg, QMessageBox.Ok)
                        return False

                    if not self.cb_info_leave.isChecked():
                        query = "INSERT INTO staff_worker_login (Worker_Info_Id, Login, Password) VALUES (%s, %s, %s)"
                        parametrs = (self.id_info, self.id_info, "")
                        info_sql = my_sql.sql_change(query, parametrs)
                        if "mysql.connector.errors" in str(type(info_sql)):
                            QMessageBox.critical(self, "Ошибка sql", info_sql.msg, QMessageBox.Ok)
                            return False

                if "passport" in self.alert and "passport" in self.alert2:
                    query = "INSERT INTO staff_worker_passport (Worker_Info_Id, Series, Number, Issued, Data_Issued, Date_Ending) VALUES (%s, %s, %s, %s, %s, %s)"
                    parametrs = (self.id_info, self.le_passport_series.text(), self.le_passport_number.text(), self.le_passport_issued.text(),
                                 self.de_passport_issued.date().toString(Qt.ISODate), self.de_passport_ending.date().toString(Qt.ISODate))
                    info_sql = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql", info_sql.msg, QMessageBox.Ok)
                        return False

                if "migration" in self.alert and "migration" in self.alert2:
                    query = "INSERT INTO staff_worker_migration (Worker_Info_Id, Serial, Number, KPP, Date_migration, Date_Validity_From, Date_Validity_To) VALUES (%s, %s, %s, %s, %s, %s, %s)"
                    parametrs = (self.id_info, self.le_migration_serial.text(), self.le_migration_number.text(), self.le_migration_kpp.text(),
                                 self.de_migration.date().toString(Qt.ISODate), self.de_migration_validity_from.date().toString(Qt.ISODate),
                                 self.de_migration_validity_to.date().toString(Qt.ISODate))
                    info_sql = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql", info_sql.msg, QMessageBox.Ok)
                        return False

                if "registration" in self.alert and "registration" in self.alert2:
                    query = "INSERT INTO staff_worker_registraton (Worker_Info_id, Address, Date_Registration, Date_Validity_From, Date_Validity_To) VALUES (%s, %s, %s, %s, %s)"
                    parametrs = (self.id_info, self.le_registration_address.text(), self.de_registration.date().toString(Qt.ISODate),
                                 self.de_registration_validity_from.date().toString(Qt.ISODate), self.de_registration_validity_to.date().toString(Qt.ISODate))
                    info_sql = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql", info_sql.msg, QMessageBox.Ok)
                        return False

                if "patent" in self.alert and "patent" in self.alert2:
                    query = "INSERT INTO staff_worker_patent (Worker_Info_Id, Serial, Number, Additional_Number, Issued, Data_Issued, Date_Ending) VALUES (%s, %s, %s, %s, %s, %s, %s)"
                    parametrs = (self.id_info, self.le_patent_serial.text(), self.le_patent_number.text(), self.le_patent_additional_number.text(),
                                 self.le_patent_issued.text(), self.de_patent_issued.date().toString(Qt.ISODate),
                                 self.de_patent_ending.date().toString(Qt.ISODate))
                    info_sql = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql", info_sql.msg, QMessageBox.Ok)
                        return False

                if "insurance" in self.alert and "insurance" in self.alert2:
                    query = "INSERT INTO staff_worker_insurance (Worker_Info_Id, Number, Company, Date) VALUES (%s, %s, %s, %s)"
                    parametrs = (self.id_info, self.le_insurance_number.text(), self.le_insurance_company.text(), self.de_insurance_date.date().toString(Qt.ISODate))
                    info_sql = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql", info_sql.msg, QMessageBox.Ok)
                        return False

                if "notification" in self.alert:
                    query = "INSERT INTO staff_worker_notification (Worker_Info_Id, Note, Date) VALUES (%s, %s, %s)"
                    parametrs = (self.id_info, self.le_notofication.toPlainText(), self.de_notification.date().toString(Qt.ISODate))
                    info_sql = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql", info_sql.msg, QMessageBox.Ok)
                        return False

                self.alert = []  # Обнуляем массив для запоминания изменений
                self.close()
                self.m.set_info()
                self.destroy()

    def exel_in(self):
        self.build_new_exel("in")

    def exel_out(self):
        self.build_new_exel("out")

    def exel_petition_in(self):
        self.build_word_petition("in")

    def exel_petition_out(self):
        self.build_word_petition("out")

    # Уведомление о заключении трудового договора
    def build_new_exel(self, option):
        # Уведомление о заключении трудового договора
        # row = (17, 30, 32, 34, 36, 38, 40, 42, 44, 46, 48, 50, 53, 55, 57, 59, 61, 63, 65)
        # col = ("A", "D", "G", "j", "M", "P", "S", "V", "Y", "AB", "AE", "AH", "AK", "AN", "AQ", "AT", "AW", "AZ", "BC", "BF", "BI", "BL", "BO", "BR", "BU", "BX",
        #       "CA", "CD", "CG", "CJ", "CM", "CP", "CS", "CV", "CY", "DB", "DE", "DH", "DK", "DN", "DQ", "DT", "DW")

        to_excel.patch_worksheet()

        self.statusBar().showMessage("Открываю шаблон")
        if option == "in":
            info = InfoDate(self.de_info_recruitment.date())
            if info.exec() == 0:
                return False
            book = openpyxl.load_workbook(filename=getcwd() + '/templates/staff/notif_in.xlsx')
        elif option == "out":
            book = openpyxl.load_workbook(filename=getcwd() + '/templates/staff/notif_out.xlsx')
        sheet = book['s1']

        col = ("Y", "AB", "AE", "AH", "AK", "AN", "AQ", "AT", "AW", "AZ", "BC", "BF", "BI", "BL", "BO", "BR", "BU", "BX",
               "CA", "CD", "CG", "CJ", "CM", "CP", "CS", "CV", "CY", "DB", "DE", "DH", "DK", "DN", "DQ", "DT", "DW")

        text = tuple(self.le_info_last_name.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Фамилия длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 53)] = t
            i += 1

        text = tuple(self.le_info_first_name.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Имя длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 55)] = t
            i += 1

        text = tuple(self.le_info_middle_name.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Отчество длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 57)] = t
            i += 1

        text = tuple(self.cb_info_country.currentText().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Место рождения длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 59)] = t
            i += 1

        # col = ("Y61", "AB61", "AE61", "AH61", "AK61", "AN61", "AQ61", "AT61", "AW61", "AZ61", "BC61", "BF61", "BI61", "BL61", "BO61", "BR61", "BU61", "BX61", "CA61",
        #        "CD61", "CG61", "CJ61", "CM61", "CP61", "CS61", "CV61", "CY61", "DB61", "DE61", "DH61", "DK61", "DN61", "DQ61", "DT61", "DW61", "A63", "D63", "G63",
        #        "j63", "M63", "P63", "S63", "V63", "Y63", "AB63", "AE63", "AH63", "AK63", "AN63", "AQ63", "AT63", "AW63", "AZ63", "BC63", "BF63", "BI63", "BL63",
        #        "BO63", "BR63", "BU63", "BX63", "CA63", "CD63", "CG63", "CJ63", "CM63", "CP63", "CS63", "CV63", "CY63", "DB63", "DE63", "DH63", "DK63", "DN63",
        #        "DQ63", "DT63", "DW63")
        col = ("Y61", "AB61", "AE61", "AH61", "AK61", "AN61", "AQ61", "AT61", "AW61", "AZ61", "BC61", "BF61", "BI61", "BL61", "BO61", "BR61", "BU61", "BX61", "CA61",
               "CD61", "CG61", "CJ61", "CM61", "CP61", "CS61", "CV61", "CY61", "DB61", "DE61", "DH61", "DK61", "DN61", "DQ61", "DT61", "DW61")
        text = tuple(self.le_info_birthplace.text().upper())
        n = 0
        if len(col) < len(text):
            m = len(col)
            while text[m] != " ":
                m -= 1
        else:
            m = len(text)
        i = 0
        while n < m:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s' % col[i]] = text[n]
            n += 1
            i += 1

        col = ("A63", "D63", "G63", "j63", "M63", "P63", "S63", "V63", "Y63", "AB63", "AE63", "AH63", "AK63", "AN63", "AQ63", "AT63", "AW63", "AZ63", "BC63", "BF63",
               "BI63", "BL63", "BO63", "BR63", "BU63", "BX63", "CA63", "CD63", "CG63", "CJ63", "CM63", "CP63", "CS63", "CV63", "CY63", "DB63", "DE63", "DH63",
               "DK63", "DN63", "DQ63", "DT63", "DW63")
        if m < len(text):
            if len(text) - m > len(col):
                QMessageBox.critical(self, "Ошибка", "Место рождения длиннее строки ввода", QMessageBox.Ok)
                return False
            else:
                m = len(text)
        i = 0
        while n < m:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s' % col[i]] = text[n]
            n += 1
            i += 1

        col = ("V", "Y", "AH", "AK", "AQ", "AT", "AW", "AZ")
        text = tuple(self.de_info_birth.date().toString("ddMMyyyy"))
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 65)] = t
            i += 1

        if self.rb_sex_m.isChecked():  # Узнаем пол работника
            sheet['%s%s' % ("BL", 65)] = "Х"
        elif self.rb_sex_f.isChecked():
            sheet['%s%s' % ("BU", 65)] = "Х"

        # Заполняем пасспорт
        sheet = book['s2']
        col = ("M", "P", "S", "V", "Y", "AB", "AE", "AH")
        text = tuple(self.le_passport_series.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Серия паспорта длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 5)] = t
            i += 1

        col = ("AT", "AW", "AZ", "BC", "BF", "BI", "BL", "BO", "BR", "BU", "BX")
        text = tuple(self.le_passport_number.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Номер паспорта длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 5)] = t
            i += 1

        col = ("CP", "CS", "DB", "DE", "DN", "DQ", "DT", "DW")
        text = tuple(self.de_passport_issued.date().toString("ddMMyyyy"))
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Дата выдачи паспорта длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 5)] = t
            i += 1

        col = ("P", "S", "V", "Y", "AB", "AE", "AH", "AK", "AN", "AQ", "AT", "AW", "AZ", "BC", "BF", "BI", "BL", "BO", "BR", "BU", "BX", "CA", "CD", "CG", "CJ",
               "CM", "CP", "CS", "CV", "CY", "DB", "DE", "DH", "DK", "DN", "DQ", "DT", "DW")
        text = tuple(self.le_passport_issued.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Кем выдан паспорт, длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 7)] = t
            i += 1

        # Миграционка
        col = ("AN", "AQ", "AT", "AW", "AZ", "BC", "BF", "BI", "BL", "BO", "BR", "BU", "BX")
        text = tuple(self.le_migration_number.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Номер миграционки длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 9)] = t
            i += 1

        # Регистрация
        col = ("CP", "CS", "DB", "DE", "DN", "DQ", "DT", "DW")
        text = tuple(self.de_migration.date().toString("ddMMyyyy"))
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Дата выдачи миграционки длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 9)] = t
            i += 1

        col = (
        "AW12", "AZ12", "BC12", "BF12", "BI12", "BL12", "BO12", "BR12", "BU12", "BX12", "CA12", "CD12", "CG12", "CJ12", "CM12", "CP12", "CS12", "CV12", "CY12",
        "DB12", "DE12", "DH12", "DK12", "DN12", "DQ12", "DT12", "DW12")
        text = tuple(self.le_registration_address.text().upper())
        n = 0
        if len(col) < len(text):
            m = len(col)
            while text[m] != " ":
                m -= 1
        else:
            m = len(text)
        i = 0
        while n < m:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s' % col[i]] = text[n]
            n += 1
            i += 1

        col = ("A14", "D14", "G14", "j14", "M14", "P14", "S14", "V14", "Y14", "AB14", "AE14", "AH14", "AK14", "AN14", "AQ14", "AT14", "AW14", "AZ14", "BC14",
               "BF14", "BI14", "BL14", "BO14", "BR14", "BU14", "BX14", "CA14", "CD14", "CG14", "CJ14", "CM14", "CP14", "CS14", "CV14", "CY14", "DB14", "DE14",
               "DH14", "DK14", "DN14", "DQ14", "DT14", "DW14")
        if m < len(text):
            if len(text) - m > len(col):
                m += len(col)
                while text[m] != " ":
                    m -= 1
            else:
                m = len(text)
        i = 0
        while n < m:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s' % col[i]] = text[n]
            n += 1
            i += 1

        col = ("A16", "D16", "G16", "j16", "M16", "P16", "S16", "V16", "Y16", "AB16", "AE16", "AH16", "AK16", "AN16", "AQ16", "AT16", "AW16", "AZ16", "BC16", "BF16",
               "BI16", "BL16", "BO16", "BR16", "BU16", "BX16", "CA16", "CD16", "CG16", "CJ16", "CM16", "CP16", "CS16", "CV16", "CY16", "DB16", "DE16", "DH16",
               "DK16", "DN16", "DQ16", "DT16", "DW16")
        if m < len(text):
            if len(text) - m > len(col):
                QMessageBox.critical(self, "Ошибка", "Место регистрации длиннее строки ввода", QMessageBox.Ok)
                return False
            else:
                m = len(text)
        i = 0
        while n < m:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s' % col[i]] = text[n]
            n += 1
            i += 1

        col = ("AW", "AZ", "BI", "BL", "BU", "BX", "CA", "CD")
        text = tuple(self.de_registration.date().toString("ddMMyyyy"))
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Дата регистрации длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 18)] = t
            i += 1

        # Патент
        country = my_sql.sql_select("SELECT Patent, Act FROM staff_country WHERE Country_name = %s", (self.cb_info_country.currentText(),))[0]
        if country[0] == 1:
            col = ("AB", "AE", "AH", "AK", "AN", "AQ", "AT", "AW", "AZ", "BC", "BF", "BI", "BL", "BO", "BR", "BU", "BX",
                   "CA", "CD", "CG", "CJ", "CM", "CP", "CS", "CV", "CY", "DB", "DE", "DH", "DK", "DN", "DQ", "DT", "DW")
            text = tuple("ПАТЕНТ " + self.le_patent_additional_number.text().upper())
            if len(text) > len(col):
                QMessageBox.critical(self, "Ошибка", "Доп номер патента длиннее строки ввода", QMessageBox.Ok)
                return False
            i = 0
            for t in text:
                self.statusBar().showMessage("Создаю %s" % i)
                sheet['%s%s' % (col[i], 22)] = t
                i += 1

            col = ("P", "S", "V", "Y", "AB", "AE", "AH", "AK", "AN", "AQ", "AT", "AW")
            text = tuple(self.le_patent_serial.text().upper())
            if len(text) > len(col):
                QMessageBox.critical(self, "Ошибка", "Серия патента длиннее строки ввода", QMessageBox.Ok)
                return False
            i = 0
            for t in text:
                self.statusBar().showMessage("Создаю %s" % i)
                sheet['%s%s' % (col[i], 24)] = t
                i += 1

            col = ("BF", "BI", "BL", "BO", "BR", "BU", "BX", "CA", "CD", "CG", "CJ", "CM")
            text = tuple(self.le_patent_number.text().upper())
            if len(text) > len(col):
                QMessageBox.critical(self, "Ошибка", "Номер патента длиннее строки ввода", QMessageBox.Ok)
                return False
            i = 0
            for t in text:
                self.statusBar().showMessage("Создаю %s" % i)
                sheet['%s%s' % (col[i], 24)] = t
                i += 1

            col = ("P", "S", "AB", "AE", "AN", "AQ", "AT", "AW")
            text = tuple(self.de_patent_issued.date().toString("ddMMyyyy"))
            if len(text) > len(col):
                QMessageBox.critical(self, "Ошибка", "Дата выдачи патента длиннее строки ввода", QMessageBox.Ok)
                return False
            i = 0
            for t in text:
                self.statusBar().showMessage("Создаю %s" % i)
                sheet['%s%s' % (col[i], 26)] = t
                i += 1

            col = ("P", "S", "V", "Y", "AB", "AE", "AH", "AK", "AN", "AQ", "AT", "AW", "AZ", "BC", "BF", "BI", "BL", "BO", "BR", "BU", "BX",
                   "CA", "CD", "CG", "CJ", "CM", "CP", "CS", "CV", "CY", "DB", "DE", "DH", "DK", "DN", "DQ", "DT", "DW")
            text = tuple(self.le_patent_issued.text().upper())
            if len(text) > len(col):
                QMessageBox.critical(self, "Ошибка", "Кем выдан патент длиннее строки ввода", QMessageBox.Ok)
                return False
            i = 0
            for t in text:
                self.statusBar().showMessage("Создаю %s" % i)
                sheet['%s%s' % (col[i], 28)] = t
                i += 1

            col = ("S", "V", "AE", "AH", "AQ", "AT", "AW", "AZ")
            text = tuple(self.de_patent_issued.date().toString("ddMMyyyy"))
            if len(text) > len(col):
                QMessageBox.critical(self, "Ошибка", "Срок С патента длиннее строки ввода", QMessageBox.Ok)
                return False
            i = 0
            for t in text:
                self.statusBar().showMessage("Создаю %s" % i)
                sheet['%s%s' % (col[i], 30)] = t
                i += 1

            col = ("BL", "BO", "BX", "CA", "CJ", "CM", "CP", "CS")
            text = tuple(self.de_patent_ending.date().toString("ddMMyyyy"))
            if len(text) > len(col):
                QMessageBox.critical(self, "Ошибка", "Срок ДО патента длиннее строки ввода", QMessageBox.Ok)
                return False
            i = 0
            for t in text:
                self.statusBar().showMessage("Создаю %s" % i)
                sheet['%s%s' % (col[i], 30)] = t
                i += 1
        else:
            text = tuple(country[1].upper())
            col = (("A33", "D33", "G33", "j33", "M33", "P33", "S33", "V33", "Y33", "AB33", "AE33", "AH33", "AK33", "AN33", "AQ33", "AT33", "AW33", "AZ33", "BC33",
                    "BF33", "BI33", "BL33", "BO33", "BR33", "BU33", "BX33", "CA33", "CD33", "CG33", "CJ33", "CM33", "CP33", "CS33", "CV33", "CY33", "DB33", "DE33",
                    "DH33", "DK33", "DN33", "DQ33", "DT33", "DW33"),
                   ("A35", "D35", "G35", "j35", "M35", "P35", "S35", "V35", "Y35", "AB35", "AE35", "AH35", "AK35", "AN35", "AQ35", "AT35", "AW35", "AZ35", "BC35",
                    "BF35", "BI35", "BL35", "BO35", "BR35", "BU35", "BX35", "CA35", "CD35", "CG35", "CJ35", "CM35", "CP35", "CS35", "CV35", "CY35", "DB35", "DE35",
                    "DH35", "DK35", "DN35", "DQ35", "DT35", "DW35"),
                   ("A37", "D37", "G37", "j37", "M37", "P37", "S37", "V37", "Y37", "AB37", "AE37", "AH37", "AK37", "AN37", "AQ37", "AT37", "AW37", "AZ37", "BC37",
                    "BF37", "BI37", "BL37", "BO37", "BR37", "BU37", "BX37", "CA37", "CD37", "CG37", "CJ37", "CM37", "CP37", "CS37", "CV37", "CY37", "DB37", "DE37",
                    "DH37", "DK37", "DN37", "DQ37", "DT37", "DW37"),
                   ("A39", "D39", "G39", "j39", "M39", "P39", "S39", "V39", "Y39", "AB39", "AE39", "AH39", "AK39", "AN39", "AQ39", "AT39", "AW39", "AZ39", "BC39",
                    "BF39", "BI39", "BL39", "BO39", "BR39", "BU39", "BX39", "CA39", "CD39", "CG39", "CJ39", "CM39", "CP39", "CS39", "CV39", "CY39", "DB39", "DE39",
                    "DH39", "DK39", "DN39", "DQ39", "DT39", "DW39"),
                   ("A41", "D41", "G41", "j41", "M41", "P41", "S41", "V41", "Y41", "AB41", "AE41", "AH41", "AK41", "AN41", "AQ41", "AT41", "AW41", "AZ41", "BC41",
                    "BF41", "BI41", "BL41", "BO41", "BR41", "BU41", "BX41", "CA41", "CD41", "CG41", "CJ41", "CM41", "CP41", "CS41", "CV41", "CY41", "DB41", "DE41",
                    "DH41", "DK41", "DN41", "DQ41", "DT41", "DW41"))
            m, n = 0, 0
            for t in text:
                if t == "\n":
                    m += 1
                    n = 0
                else:
                    self.statusBar().showMessage("Создаю %s" % n)
                    sheet['%s' % (col[m][n])] = t
                    n += 1

        # Должность
        position_number = my_sql.sql_select("SELECT Number FROM staff_position WHERE Name = %s", (self.cb_info_position.currentText(),))[0][0]
        col = ("A", "D", "G", "j", "M", "P", "S", "V", "Y", "AB", "AE", "AH", "AK", "AN", "AQ", "AT", "AW", "AZ", "BC", "BF", "BI", "BL", "BO", "BR", "BU", "BX",
               "CA", "CD", "CG", "CJ", "CM", "CP", "CS", "CV", "CY", "DB", "DE", "DH", "DK", "DN", "DQ", "DT", "DW")
        text = tuple(self.cb_info_position.currentText().upper() + " " + position_number)
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Должность длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 44)] = t
            i += 1

        # Дата работы / увольнения
        col = ("CP", "CS", "DB", "DE", "DN", "DQ", "DT", "DW")
        if option == "in":
            text = tuple(info.de_in.date().toString("ddMMyyyy"))
        elif option == "out":
            text = tuple(self.de_info_leave.date().toString("ddMMyyyy"))
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "дата договора о работе патента длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 50)] = t
            i += 1

        dir_name = self.le_info_last_name.text() + " " + self.le_info_first_name.text() + " " + self.de_info_recruitment.date().toString("dd.MM.yyyy")
        self.path = self.inspection_path(dir_name, 'Путь корень рабочие')
        if self.path:
            self.statusBar().showMessage("Сохраняю")
            if option == "in":
                book.save('%s/%s' % (self.path, "Уведомление о приеме на работу.xlsx"))
            elif option == "out":
                book.save('%s/%s' % (self.path, "Уведомление о увольнении.xlsx"))
            dir_name = self.le_info_last_name.text() + " " + self.le_info_first_name.text() + " " + self.de_info_recruitment.date().toString("dd.MM.yyyy")
            self.path = self.inspection_path(dir_name, 'Путь корень рабочие')
            self.statusBar().showMessage("Готово")
        else:
            self.statusBar().showMessage("Ошибка сохранения")
            return False

    # Уведомление о регистрации
    def build_exel_registration(self):

        to_excel.patch_worksheet()

        info = ExelInfo(self.le_info_birthplace.text())
        if info.exec() == 0:
            return False
        self.statusBar().showMessage("Открываю шаблон")
        book = openpyxl.load_workbook(filename=getcwd() + '/templates/staff/registration.xlsx')
        sheet = book["s1"]
        col = ("N", "Q", "T", "W", "Z", "AC", "AF", "AI", "AL", "AO", "AR", "AU", "AX", "BA", "BD", "BG", "BJ", "BM", "BP", "BS", "BV", "BY", "CB", "CE", "CH", "CK",
               "CN", "CQ", "CT", "CW", "CZ", "DC", "DF", "DI", "DL")

        text = tuple(self.le_info_last_name.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Фамилия длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 16)] = t
            sheet['%s%s' % (col[i], 69)] = t
            i += 1

        text = tuple(self.le_info_first_name.text().upper() + " " + self.le_info_middle_name.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Имя + отчество длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 18)] = t
            sheet['%s%s' % (col[i], 71)] = t
            i += 1

        col = ("Q", "T", "W", "Z", "AC", "AF", "AI", "AL", "AO", "AR", "AU", "AX", "BA", "BD", "BG", "BJ", "BM", "BP", "BS", "BV", "BY", "CB", "CE", "CH", "CK",
               "CN", "CQ", "CT", "CW", "CZ", "DC", "DF", "DI", "DL")
        text = tuple(self.cb_info_country.currentText().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Гражданство длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 21)] = t
            sheet['%s%s' % (col[i], 74)] = t
            i += 1

        col = ("T", "W", "AF", "AI", "AO", "AR", "AU", "AX")
        text = tuple(self.de_info_birth.date().toString("ddMMyyyy"))
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 24)] = t
            sheet['%s%s' % (col[i], 77)] = t
            i += 1

        if self.rb_sex_m.isChecked():  # Узнаем пол работника
            sheet['%s%s' % ("BV", 24)] = "Х"
            sheet['%s%s' % ("BY", 77)] = "Х"
        elif self.rb_sex_f.isChecked():
            sheet['%s%s' % ("CK", 24)] = "Х"
            sheet['%s%s' % ("CN", 77)] = "Х"

        col = ("T", "W", "Z", "AC", "AF", "AI", "AL", "AO", "AR", "AU", "AX", "BA", "BD", "BG", "BJ", "BM", "BP", "BS", "BV", "BY", "CB", "CE", "CH", "CK",
               "CN", "CQ", "CT", "CW", "CZ", "DC", "DF", "DI", "DL")
        text = tuple(info.le_birthplace_country.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Государство места рождения длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 27)] = t
            i += 1

        text = tuple(info.le_birthplace_city.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Город места рождения длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 30)] = t
            i += 1

        col = ("BY", "CB", "CE", "CH")
        text = tuple(self.le_passport_series.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Серия паспорта длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 33)] = t
            sheet['%s%s' % (col[i], 80)] = t
            i += 1

        col = ("CN", "CQ", "CT", "CW", "CZ", "DC", "DF", "DI", "DL")
        text = tuple(self.le_passport_number.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Номер паспорта длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 33)] = t
            sheet['%s%s' % (col[i], 80)] = t
            i += 1

        col = ("Q", "T", "AC", "AF", "AL", "AO", "AR", "AU")
        text = tuple(self.de_passport_issued.date().toString("ddMMyyyy"))
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 35)] = t
            i += 1

        col = ("BM", "BP", "BY", "CB", "CH", "CK", "CN", "CQ")
        text = tuple(self.de_passport_ending.date().toString("ddMMyyyy"))
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 35)] = t
            i += 1

        col = ("N", "Q", "T", "W", "Z", "AC", "AF", "AI", "AL", "AO", "AR", "AU", "AX", "BA", "BD", "BG", "BJ", "BM", "BP", "BS", "BV", "BY", "CB")
        text = tuple(self.cb_info_position.currentText().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Должность длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 46)] = t
            i += 1

        col = ("W", "Z", "AI", "AL", "AR", "AU", "AX", "BA")
        text = tuple(info.de_in.date().toString("ddMMyyyy"))
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 48)] = t
            i += 1

        col = ("CH", "CK", "CT", "CW", "DC", "DF", "DI", "DL")
        text = tuple(info.de_from.date().toString("ddMMyyyy"))
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 48)] = t
            i += 1

        col = ("AC", "AF", "AO", "AR", "AX", "BA", "BD", "BG")
        text = tuple(info.de_from.date().toString("ddMMyyyy"))
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 94)] = t
            i += 1

        col = ("AC", "AF", "AI", "AL")
        text = tuple(self.le_migration_serial.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Серия миграционки длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 50)] = t
            i += 1

        col = ("AR", "AU", "AX", "BA", "BD", "BG", "BJ", "BM")
        text = tuple(self.le_migration_number.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Номер миграционки длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 50)] = t
            i += 1

        # Вставляем черные квадраты
        img = Image('%s/staff/square.png' % self.path_templates)
        sheet.add_image(img, 'E7')
        img = Image('%s/staff/square.png' % self.path_templates)
        sheet.add_image(img, 'DJ7')
        img = Image('%s/staff/square.png' % self.path_templates)
        sheet.add_image(img, 'E96')
        img = Image('%s/staff/square.png' % self.path_templates)
        sheet.add_image(img, 'CO96')
        img = Image('%s/staff/square.png' % self.path_templates)
        sheet.add_image(img, 'DJ96')

        sheet = book["s2"]
        img = Image('%s/staff/square.png' % self.path_templates)
        sheet.add_image(img, 'E3')
        img = Image('%s/staff/square.png' % self.path_templates)
        sheet.add_image(img, 'DJ3')
        img = Image('%s/staff/square.png' % self.path_templates)
        sheet.add_image(img, 'E79')
        img = Image('%s/staff/square.png' % self.path_templates)
        sheet.add_image(img, 'CO79')
        img = Image('%s/staff/square.png' % self.path_templates)
        sheet.add_image(img, 'DJ79')

        dir_name = self.le_info_last_name.text() + " " + self.le_info_first_name.text() + " " + self.de_info_recruitment.date().toString("dd.MM.yyyy")
        self.path = self.inspection_path(dir_name, 'Путь корень рабочие')
        if self.path:
            self.statusBar().showMessage("Сохраняю фаил")
            book.save('%s/%s' % (self.path, "Уведомление для регистрации.xlsx"))
            self.inspection_files(dir_name, 'Путь корень рабочие')
            self.statusBar().showMessage("Готово")
        else:
            self.statusBar().showMessage("Ошибка сохранения")
            return False

    # Договор на работу
    def build_word_in(self):

        if not self.id_info:
            QMessageBox.critical(self, "Ошибка", "У этого работника нет номера", QMessageBox.Ok)
            return False

        # Проверяем нужный номер документа
        self.statusBar().showMessage("проверяю SQL")
        doc_date_new = False
        query = "SELECT IFNULL(MAX(Number), 'No Number'), Date FROM staff_worker_doc_number WHERE Worker_Info_Id = %s AND Name = %s"
        doc_number_sql = my_sql.sql_select(query, (self.id_info, "труд.дог."))
        if "mysql.connector.errors" in str(type(doc_number_sql)):
            QMessageBox.critical(self, "Ошибка sql", doc_number_sql.msg, QMessageBox.Ok)

        if "No Number" in doc_number_sql[0]:
            doc_date_new = True
            query = "SELECT IFNULL(MAX(Number)+1, 'No Number'), NOW() FROM staff_worker_doc_number WHERE YEAR(Date) = %s AND Name = %s"
            doc_number_sql = my_sql.sql_select(query, (QDate.currentDate().year(), "труд.дог."))
            if "mysql.connector.errors" in str(type(doc_number_sql)):
                QMessageBox.critical(self, "Ошибка sql", doc_number_sql.msg, QMessageBox.Ok)

        if "No Number" not in doc_number_sql[0]:
            doc_number = doc_number_sql[0][0]
            doc_date = doc_number_sql[0][1]
        else:
            doc_number = 1
            doc_date = self.de_info_recruitment.date().toPyDate()

        info = InfoDate(doc_date)
        if info.exec() == 0:
            self.statusBar().showMessage("Отмена")
            return False

        if info.de_in.date().toPyDate() != doc_date:
            doc_date = info.de_in.date().toPyDate()
            doc_date_new = True

        # Нужен ли патент
        patent = my_sql.sql_select("SELECT Patent FROM staff_country WHERE Country_name = %s", (self.cb_info_country.currentText(),))[0][0]

        self.statusBar().showMessage("Открываю шаблон")
        f = open(getcwd() + '/templates/staff/contract.xml', "r", -1, "utf-8")
        xml = f.read()
        self.statusBar().showMessage("Закрываю шаблон")
        f.close()
        self.statusBar().showMessage("Создаю документ")
        if patent == 1:
            number_xml = "0" + str(doc_number) + "/" + doc_date.strftime("%y")
        else:
            number_xml = str(doc_number) + "/" + doc_date.strftime("%y")

        xml = xml.replace("НОМЕР", number_xml)
        xml = xml.replace("ДАТА", info.de_in.date().toString("dd.MM.yyyy"))
        xml = xml.replace("ФИО", self.le_info_last_name.text() + " " + self.le_info_first_name.text() + " " + self.le_info_middle_name.text())
        if self.rb_sex_m.isChecked():  # Узнаем пол работника
            xml = xml.replace("ЫЙ/АЯ", "ый")
        elif self.rb_sex_f.isChecked():
            xml = xml.replace("ЫЙ/АЯ", "ая")
        xml = xml.replace("ПРОФФ", self.cb_info_position.currentText().lower())
        if patent == 1:
            xml = xml.replace("ВРЕМЯ", "время действия патента серии")
            xml = xml.replace("ПАТЕНТ", self.le_patent_serial.text() + " № " + self.le_patent_number.text() + ".")
        else:
            xml = xml.replace("ВРЕМЯ", "неопределенный срок.")
            xml = xml.replace("ПАТЕНТ", "")
        xml = xml.replace("ДАТРОЖ", self.de_info_birth.date().toString("dd.MM.yyyy"))
        xml = xml.replace("ГРАЖДАНСТВО", self.cb_info_country.currentText())
        xml = xml.replace("ПАССПОРТ", self.le_passport_series.text().upper() + " " + self.le_passport_number.text().upper())
        if patent == 1:
            xml = xml.replace("?ПА", "Патент: ")
            xml = xml.replace("СТРАХОВКА", "ДМС: №" + self.le_insurance_number.text().upper() + ", от " + self.de_insurance_date.date().toString("dd.MM.yyyy") +
                              "г., страховая компания \"" + self.le_insurance_company.text() + '\"')
        else:
            xml = xml.replace("СТРАХОВКА", "")
            xml = xml.replace("?ПА", "")

        dir_name = self.le_info_last_name.text() + " " + self.le_info_first_name.text() + " " + self.de_info_recruitment.date().toString("dd.MM.yyyy")
        self.inspection_files(dir_name, 'Путь корень рабочие')
        if self.path:
            self.statusBar().showMessage("Сохраняю фаил")
            f = open('%s/%s' % (self.path, "Трудовой договор.doc"), "w", -1, "utf-8")
            f.write(xml)
            f.close()
            if doc_date_new:
                query = """INSERT INTO staff_worker_doc_number (Worker_Info_Id, Name, Number, Date) VALUES (%s, %s, %s, %s)
                            ON DUPLICATE KEY UPDATE Date = %s"""
                parametrs = (self.id_info, "труд.дог.", doc_number, doc_date, doc_date)
                info_sql = my_sql.sql_change(query, parametrs)
                if "mysql.connector.errors" in str(type(info_sql)):
                    QMessageBox.critical(self, "Ошибка sql", info_sql.msg, QMessageBox.Ok)
                    return False
            self.statusBar().showMessage("Готово")
            self.inspection_files(dir_name, 'Путь корень рабочие')
        else:
            self.statusBar().showMessage("Ошибка сохранения")
            return False

    # Заявления на прием и увольнение
    def build_word_petition(self, option):

        if option == "out":
            info = InfoDate(QDate().currentDate())
            info.label.setText("Дата увольнения")
            if info.exec() == 0:
                return False

        if not self.id_info:
            QMessageBox.critical(self, "Ошибка", "У этого работника нет номера", QMessageBox.Ok)
            return False

        self.statusBar().showMessage("Открываю шаблон")
        if option == "out":
            f = open(getcwd() + '/templates/staff/petition_out.xml', "r", -1, "utf-8")
        else:
            f = open(getcwd() + '/templates/staff/petition_in.xml', "r", -1, "utf-8")
        xml = f.read()
        self.statusBar().showMessage("Закрываю шаблон")
        f.close()
        self.statusBar().showMessage("Создаю документ")

        xml = xml.replace("?ФИО", self.le_info_last_name.text() + " " + self.le_info_first_name.text() + " " + self.le_info_middle_name.text())
        xml = xml.replace("?СЕРИЯ", self.le_passport_series.text().upper())
        xml = xml.replace("?НОМЕР", self.le_passport_number.text().upper())
        xml = xml.replace("?ВЫДАН", self.le_passport_issued.text())
        xml = xml.replace("?ДАТАВЫД", self.de_passport_issued.date().toString("dd.MM.yyyy"))
        xml = xml.replace("?ДАТАЗАЯВ", QDate().currentDate().toString("dd.MM.yyyy"))
        if option == "out":
            xml = xml.replace("?ДАТАУВОЛЬН", info.de_in.date().toString("dd.MM.yyyy"))
        else:
            xml = xml.replace("?ДОЛЖНОСТЬ", self.cb_info_position.currentText())

        dir_name = self.le_info_last_name.text() + " " + self.le_info_first_name.text() + " " + self.de_info_recruitment.date().toString("dd.MM.yyyy")
        self.inspection_files(dir_name, 'Путь корень рабочие')
        if self.path:
            self.statusBar().showMessage("Сохраняю фаил")
            if option == "out":
                f = open('%s/%s' % (self.path, "Заявление об увольнении.doc"), "w", -1, "utf-8")
            else:
                f = open('%s/%s' % (self.path, "Заявление о приеме на работу.doc"), "w", -1, "utf-8")
            f.write(xml)
            f.close()
            self.statusBar().showMessage("Готово")
            self.inspection_files(dir_name, 'Путь корень рабочие')
        else:
            self.statusBar().showMessage("Ошибка сохранения")
            return False

    # Довереность на ЗП
    def build_word_proxy(self):

        if not self.id_info:
            QMessageBox.critical(self, "Ошибка", "У этого работника нет номера", QMessageBox.Ok)
            return False

        self.statusBar().showMessage("Открываю шаблон")
        f = open(getcwd() + '/templates/staff/proxy.xml', "r", -1, "utf-8")
        xml = f.read()
        self.statusBar().showMessage("Закрываю шаблон")
        f.close()
        self.statusBar().showMessage("Создаю документ")

        xml = xml.replace("?ФИО", self.le_info_last_name.text() + " " + self.le_info_first_name.text() + " " + self.le_info_middle_name.text())
        xml = xml.replace("?СЕРИЯ", self.le_passport_series.text().upper())
        xml = xml.replace("?НОМЕР", self.le_passport_number.text().upper())
        xml = xml.replace("?ДАТАЗАЯВ", QDate().currentDate().toString("dd.MM.yyyy"))

        dir_name = self.le_info_last_name.text() + " " + self.le_info_first_name.text() + " " + self.de_info_recruitment.date().toString("dd.MM.yyyy")
        self.inspection_files(dir_name, 'Путь корень рабочие')
        if self.path:
            self.statusBar().showMessage("Сохраняю фаил")
            f = open('%s/%s' % (self.path, "Доверенность на ЗП.doc"), "w", -1, "utf-8")
            f.write(xml)
            f.close()
            self.statusBar().showMessage("Готово")
            self.inspection_files(dir_name, 'Путь корень рабочие')
        else:
            self.statusBar().showMessage("Ошибка сохранения")
            return False

    # Временный пропуск
    def build_word_pass(self):

        if not self.id_info:
            QMessageBox.critical(self, "Ошибка", "У этого работника нет номера", QMessageBox.Ok)
            return False

        self.statusBar().showMessage("Открываю шаблон")
        f = open(getcwd() + '/templates/staff/pass.xml', "r", -1, "utf-8")
        xml = f.read()
        self.statusBar().showMessage("Закрываю шаблон")
        f.close()
        self.statusBar().showMessage("Создаю документ")

        xml = xml.replace("?ФИО", self.le_info_last_name.text() + " " + self.le_info_first_name.text() + " " + self.le_info_middle_name.text())

        dir_name = self.le_info_last_name.text() + " " + self.le_info_first_name.text() + " " + self.de_info_recruitment.date().toString("dd.MM.yyyy")
        self.inspection_files(dir_name, 'Путь корень рабочие')
        if self.path:
            self.statusBar().showMessage("Сохраняю фаил")
            f = open('%s/%s' % (self.path, "Временный пропуск.doc"), "w", -1, "utf-8")
            f.write(xml)
            f.close()
            self.statusBar().showMessage("Готово")
            self.inspection_files(dir_name, 'Путь корень рабочие')
        else:
            self.statusBar().showMessage("Ошибка сохранения")
            return False

    # Справка о приеме уведомления
    def build_word_admission_notice(self):

        info = InfoDate(QDate().currentDate())
        info.label.setText("Дата приема уведомления")
        if info.exec() == 0:
            return False

        if not self.id_info:
            QMessageBox.critical(self, "Ошибка", "У этого работника нет номера", QMessageBox.Ok)
            return False

        self.statusBar().showMessage("Открываю шаблон")
        f = open(getcwd() + '/templates/staff/admission_notice.xml', "r", -1, "utf-8")
        xml = f.read()
        self.statusBar().showMessage("Закрываю шаблон")
        f.close()
        self.statusBar().showMessage("Создаю документ")

        xml = xml.replace("?ФИ", self.le_info_last_name.text() + " " + self.le_info_first_name.text())
        xml = xml.replace("?ДАТА", info.de_in.date().toString("dd.MM.yyyy"))

        dir_name = self.le_info_last_name.text() + " " + self.le_info_first_name.text() + " " + self.de_info_recruitment.date().toString("dd.MM.yyyy")
        self.inspection_files(dir_name, 'Путь корень рабочие')
        if self.path:
            self.statusBar().showMessage("Сохраняю фаил")
            f = open('%s/%s' % (self.path, "Справка о приеме уведомления.doc"), "w", -1, "utf-8")
            f.write(xml)
            f.close()
            self.statusBar().showMessage("Готово")
            self.inspection_files(dir_name, 'Путь корень рабочие')
        else:
            self.statusBar().showMessage("Ошибка сохранения")
            return False

    # Ходотайство о продлении регистрации
    def build_word_hodataistvo(self):

        if not self.id_info:
            QMessageBox.critical(self, "Ошибка", "У этого работника нет номера", QMessageBox.Ok)
            return False

        # Проверяем нужный номер документа
        self.statusBar().showMessage("проверяю SQL")
        query = "SELECT IFNULL(MAX(Number), 'No Number') FROM staff_worker_doc_number WHERE Name = %s"
        doc_number = my_sql.sql_select(query, ("ходатайство", ))
        if "mysql.connector.errors" in str(type(doc_number)):
            QMessageBox.critical(self, "Ошибка sql", doc_number.msg, QMessageBox.Ok)
        doc_number = doc_number[0][0]

        if "No Number" not in doc_number:
            doc_number = int(doc_number) + 1
        else:
            doc_number = 1

        # Узнаем номер договора
        query = "SELECT IFNULL(MAX(Number), 'No Number'), Date FROM staff_worker_doc_number WHERE Name = %s AND Worker_Info_Id = %s"
        contract_number = my_sql.sql_select(query, ("труд.дог.", self.id_info))
        if "mysql.connector.errors" in str(type(doc_number)):
            QMessageBox.critical(self, "Ошибка sql", contract_number.msg, QMessageBox.Ok)
            return False

        # Нужен ли патент
        patent = my_sql.sql_select("SELECT Patent FROM staff_country WHERE Country_name = %s", (self.cb_info_country.currentText(),))[0][0]

        if not self.id_info:
            QMessageBox.critical(self, "Ошибка", "У этого работника нет номера", QMessageBox.Ok)
            return False

        self.statusBar().showMessage("Открываю шаблон")
        f = open(getcwd() + '/templates/staff/hodataistvo.xml', "r", -1, "utf-8")
        xml = f.read()
        self.statusBar().showMessage("Закрываю шаблон")
        f.close()
        self.statusBar().showMessage("Создаю документ")

        xml = xml.replace("?НОМЕРДОК", str(doc_number))
        xml = xml.replace("?ДАТАДОК", QDate().currentDate().toString("dd.MM.yyyy"))
        xml = xml.replace("?РЕСПУБЛИКА", self.cb_info_country.currentText())
        xml = xml.replace("?ФИО", self.le_info_last_name.text() + " " + self.le_info_first_name.text() + " " + self.le_info_middle_name.text())
        xml = xml.replace("?ДАТАРОЖ", self.de_info_birth.date().toString("dd.MM.yyyy"))
        xml = xml.replace("?ПАССПОРТ", self.le_passport_series.text().upper() + " " + self.le_passport_number.text().upper())
        xml = xml.replace("?ДАТАВЬЕЗД", self.de_migration.text())
        if contract_number[0][0] != 'No Number':
            xml = xml.replace("?ТРУДДОГ", "№ " + contract_number[0][0] + "/" + contract_number[0][1].strftime("%y") + " от " + contract_number[0][1].strftime("%d.%m.%Y"))
        else:
            xml = xml.replace("?ТРУДДОГ", "договор не найден")
        xml = xml.replace("?РЕГДО", self.de_registration_validity_to.date().toString("dd.MM.yyyy"))
        xml = xml.replace("?РАДР", self.le_registration_address.text())
        if patent == 1:
            xml = xml.replace("?ПАТЕНТ", "-патент на работу " + self.le_patent_serial.text() + " № " + self.le_patent_number.text() + ", действительный до " + self.de_patent_ending.date().toString("dd.MM.yyyy"))
        else:
            xml = xml.replace("?ПАТЕНТ", " ")

        dir_name = self.le_info_last_name.text() + " " + self.le_info_first_name.text() + " " + self.de_info_recruitment.date().toString("dd.MM.yyyy")
        self.inspection_files(dir_name, 'Путь корень рабочие')
        if self.path:
            self.statusBar().showMessage("Сохраняю фаил")
            f = open('%s/%s' % (self.path, "Ходатайство.doc"), "w", -1, "utf-8")
            f.write(xml)
            f.close()
            query = """INSERT INTO staff_worker_doc_number (Worker_Info_Id, Name, Number, Date) VALUES (%s, %s, %s, %s)
                        ON DUPLICATE KEY UPDATE Number = %s, Date = %s"""
            parametrs = (self.id_info, "ходатайство", doc_number, QDate.currentDate().toString(Qt.ISODate), doc_number, QDate.currentDate().toString(Qt.ISODate))
            info_sql = my_sql.sql_change(query, parametrs)
            if "mysql.connector.errors" in str(type(info_sql)):
                QMessageBox.critical(self, "Ошибка sql", info_sql.msg, QMessageBox.Ok)
                return False
            self.statusBar().showMessage("Готово")
            self.inspection_files(dir_name, 'Путь корень рабочие')
        else:
            self.statusBar().showMessage("Ошибка сохранения")
            return False

    def closeEvent(self, e):
        if self.alert:
            result = QMessageBox.question(self, "Выйтиb?", "Сохранить изменения перед выходом?",
                                          QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel, QMessageBox.Yes)
            if result == 16384:
                self.acc()
                e.accept()
            elif result == 65536:
                e.accept()
            elif result == 4194304:
                e.ignore()
        else:

            e.accept()


class StaffFilter(QDialog, staff_filter):
    def __init__(self, main):
        super(StaffFilter, self).__init__()
        self.setupUi(self)
        self.setWindowIcon(QIcon(getcwd() + "/images/icon.ico"))

        self.main = main

        # заполняем страны
        query = "SELECT Country_name, Id FROM staff_country"
        self.country = my_sql.sql_select(query)
        if "mysql.connector.errors" in str(type(self.country)):
            raise RuntimeError("Не смог получить страны")
        self.cb_info_country.addItem("", "")
        for country in self.country:
            self.cb_info_country.addItem(country[0], country[1])

    def ui_enabled_date(self, en):
        self.de_info_birth.setEnabled(en)

    def ui_acc(self):
        where = ""

        # Блок условий имя
        if self.le_info_first_name.text() != '':
            where = self.add_filter(where, "(staff_worker_info.First_Name LIKE '%s')" % ("%" + self.le_info_first_name.text() + "%", ))

        # Блок условий фамилия
        if self.le_info_last_name.text() != '':
            where = self.add_filter(where, "(staff_worker_info.Last_Name LIKE '%s')" % ("%" + self.le_info_last_name.text() + "%", ))

        # Блок условий отчество
        if self.le_info_middle_name.text() != '':
            where = self.add_filter(where, "(staff_worker_info.Middle_Name LIKE '%s')" % ("%" + self.le_info_middle_name.text() + "%", ))

        # Блок условий табельный номер
        if self.le_info_id.text() != '':
            where = self.add_filter(where, "(staff_worker_info.Id = %s)" % self.le_info_id.text())

        # Блок условий страна
        if self.cb_info_country.currentData() != '':
            where = self.add_filter(where, "(staff_worker_info.Country_Id = %s)" % self.cb_info_country.currentData())

        # Блок  условий мужчина или женщина
        where_item = ""
        if self.cb_sex_m.isChecked():
            where_item = self.add_filter(where_item, "staff_worker_info.Sex = 'M'", False)

        if self.cb_sex_f.isChecked():
            where_item = self.add_filter(where_item, "staff_worker_info.Sex = 'F'", False)

        if where_item:
            where_item = "(" + where_item + ")"
            where = self.add_filter(where, where_item)

        # Блок  условий даты рождения
        if self.cb_date.isChecked():
            where = self.add_filter(where_item, "staff_worker_info.Date_Birth = '%s'" % self.de_info_birth.date().toString(Qt.ISODate))

        # Блок  условий даты приема
        if self.gp_date_recruitment.isChecked():
            sql_date = "(staff_worker_info.Date_Recruitment >= '%s' AND staff_worker_info.Date_Recruitment <= '%s')" % \
                       (self.de_date_recruitment_from.date().toString(Qt.ISODate), self.de_date_recruitment_to.date().toString(Qt.ISODate))
            where = self.add_filter(where, sql_date)

        # Блок  условий даты увольнения
        if self.gp_date_leave.isChecked():
            sql_date = "(staff_worker_info.`Leave` = 1 AND staff_worker_info.Date_Leave >= '%s' AND staff_worker_info.Date_Leave <= '%s')" % \
                       (self.de_date_leave_from.date().toString(Qt.ISODate), self.de_date_leave_to.date().toString(Qt.ISODate))
            where = self.add_filter(where, sql_date)

        if where:
            self.sql_query_all = self.sql_query_all + " WHERE " + where

        self.main.of_set_filter(self.sql_query_all)

        self.close()

    def ui_can(self):
        self.close()
        self.destroy()

    def add_filter(self, where, add, and_add=True):
        if where:
            if and_add:
                where += " AND " + add
            else:
                where += " OR " + add
        else:
            where = add

        return where

    def of_set_sql_query(self, sql):
        self.sql_query_all = sql


class Country(list.ListItems):
    def set_settings(self):
        self.setWindowTitle("Настройка стран")  # Имя окна
        self.toolBar.setStyleSheet("background-color: rgb(129, 66, 255);")  # Цвет бара
        self.title_new_window = "Страна"  # Имя вызываемых окон

        self.sql_list = "SELECT id, staff_country.Country_name FROM staff_country ORDER BY staff_country.Country_name"
        self.sql_add = "INSERT INTO staff_country (Country_name, Patent, Act) VALUES (%s, %s, %s)"
        self.sql_change_select = "SELECT Country_name, Patent, Act FROM staff_country WHERE Id = %s"
        self.sql_update_select = 'UPDATE staff_country SET Country_name = %s, Patent = %s, Act = %s WHERE Id = %s'
        self.sql_dell = "DELETE FROM staff_country WHERE Id = %s"

        self.set_new_win = {"WinTitle": "Страна",
                            "WinColor": "(129, 66, 255)",
                            "lb_name": "Название",
                            "lb_note": "Заметка"}

    def ui_add_item(self):
        self.add_country = ChangeCountry(self)
        self.add_country.setModal(True)
        self.add_country.show()

    def ui_change_item(self, id=False):
        if id:
            id_select = id
        else:
            try:
                id_select = self.lw_list.selectedItems()[0].data(3)
            except:
                QMessageBox.critical(self, "Ошибка", "Выберете элемент", QMessageBox.Ok)
                return False

        self.change_country = ChangeCountry(self, id_select)
        self.change_country.setModal(True)
        self.change_country.show()


class ChangeCountry(QDialog, country_class):  # Ввод и изменение гражданств
    def __init__(self, main, id=None):
        super(ChangeCountry, self).__init__()
        self.setupUi(self)
        self.setWindowIcon(QIcon(getcwd() + "/images/icon.ico"))

        self.main = main
        self.id = id
        self.change_on = False

        if self.id:
            self.set_info()
            self.change_on = True
        else:
            self.change_on = False

    def set_info(self):
        self.country_id_change = self.id
        sql_ret = my_sql.sql_select(self.main.sql_change_select, (self.id, ))
        if "mysql.connector.errors" in str(type(sql_ret)):
            QMessageBox.critical(self, "Ошибка sql", sql_ret.msg, QMessageBox.Ok)
            return False
        else:
            self.le_name.setText(sql_ret[0][0])
            if not sql_ret[0][1]:
                self.cb_patent.setChecked(False)
                self.te_act.setEnabled(True)
                self.te_act.appendPlainText(sql_ret[0][2])
            else:
                self.te_act.setEnabled(False)
                self.cb_patent.setChecked(True)

    def check_patent(self, pat):
        if pat:
            self.te_act.setEnabled(False)
        else:
            self.te_act.setEnabled(True)

    def acc(self):
        name = self.le_name.text()
        if not self.cb_patent.isChecked():
            patent = 0
            act = self.te_act.toPlainText()
        else:
            patent = 1
            act = ""

        if self.change_on:
            sql_ret = my_sql.sql_change(self.main.sql_update_select, (name, patent, act, self.id))
        else:
            sql_ret = my_sql.sql_change(self.main.sql_add, (name, patent, act))
        if "mysql.connector.errors" in str(type(sql_ret)):
            QMessageBox.critical(self, "Ошибка sql", sql_ret.msg, QMessageBox.Ok)
            return False
        self.main.sql_set_list()
        self.close()
        self.destroy()

    def cancel(self):
        result = QMessageBox.question(self, "Выйти?", "Вы уверены что хотите выйти??", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if result == 16384:
            self.close()
            self.destroy()
        elif result == 65536:
            pass


class StaffPosition(list.ListItems):
    def set_settings(self):
        self.setWindowTitle("должностей")  # Имя окна
        self.toolBar.setStyleSheet("background-color: rgb(129, 66, 255);")  # Цвет бара
        self.title_new_window = "Должность"  # Имя вызываемых окон

        self.sql_list = "SELECT Id, staff_position.Name FROM staff_position"
        self.sql_add = ""
        self.sql_change_select = ""
        self.sql_update_select = ""
        self.sql_dell = "DELETE FROM staff_position WHERE Id = %s"

    def ui_add_item(self):
        self.add_position = ChangePosition(self)
        self.add_position.setModal(True)
        self.add_position.show()

    def ui_change_item(self, id=False):
        if id:
            id_select = id
        else:
            try:
                id_select = self.lw_list.selectedItems()[0].data(3)
            except:
                QMessageBox.critical(self, "Ошибка", "Выберете элемент", QMessageBox.Ok)
                return False
        self.change_provider = ChangePosition(self, id_select)
        self.change_provider.setModal(True)
        self.change_provider.show()


class ChangePosition(QDialog, position_class):
    def __init__(self, main, id=None):
        super(ChangePosition, self).__init__()
        self.main = main
        self.id = id
        self.setupUi(self)
        self.setWindowIcon(QIcon(getcwd() + "/images/icon.ico"))
        self.change_on = False

        if id:
            self.set_info()

    def set_info(self):
        self.change_on = True
        sql_ret = my_sql.sql_select("SELECT Name, Number FROM staff_position WHERE Id = %s", (self.id,))
        if "mysql.connector.errors" in str(type(sql_ret)):
            QMessageBox.critical(self, "Ошибка sql", sql_ret.msg, QMessageBox.Ok)
            return False

        self.le_name.setText(sql_ret[0][0])
        if str(sql_ret[0][1]) != "None":
            self.le_number.setText(sql_ret[0][1])

    def acc(self):
        name = self.le_name.text()
        number = self.le_number.text()
        if self.change_on:
            par = (name, number, self.id)
            sql_ret = my_sql.sql_change("UPDATE staff_position SET Name = %s, Number = %s WHERE Id = %s", par)
        else:
            par = (name, number)
            sql_ret = my_sql.sql_change("INSERT INTO staff_position(Name, Number) VALUES (%s, %s)", par)
        if "mysql.connector.errors" in str(type(sql_ret)):
            QMessageBox.critical(self, "Ошибка sql", sql_ret.msg, QMessageBox.Ok)
            return False
        self.main.sql_set_list()
        self.close()
        self.destroy()

    def cancel(self):
        result = QMessageBox.question(self, "Выйти?", "Вы уверены что хотите выйти??", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if result == 16384:
            self.close()
            self.destroy()
        elif result == 65536:
            pass


class ExelInfo(QDialog, exel_info_class):
    def __init__(self, birthplace):
        super(ExelInfo, self).__init__()
        self.setupUi(self)
        to_date = QDate.currentDate()
        self.de_in.setDate(to_date)
        self.de_from.setDate(to_date)
        self.setModal(True)
        self.setWindowIcon(QIcon(getcwd() + "/images/icon.ico"))
        self.le_birthplace_country.setText(birthplace)
        self.show()


class InfoDate(QDialog, info_date_class):
    def __init__(self, date_in):
        super(InfoDate, self).__init__()
        self.setupUi(self)
        self.de_in.setDate(date_in)
        self.setModal(True)
        self.setWindowIcon(QIcon(getcwd() + "/images/icon.ico"))
        self.show()


class AddFile(QDialog, add_file_date_class):
    def __init__(self):
        super(AddFile, self).__init__()
        self.setupUi(self)
        self.setModal(True)
        self.setWindowIcon(QIcon(getcwd() + "/images/icon.ico"))
        self.show()

    def copy_file_path(self):
        self.path_copy_file.setText(QFileDialog.getOpenFileName(self, "Выберите копируемый фаил")[0])

